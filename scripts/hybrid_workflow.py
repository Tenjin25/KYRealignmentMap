"""
Hybrid PDF Extraction Workflow for Kentucky Election Results.
Combines automated extraction with easy manual correction.

This script:
1. Extracts data from PDF automatically
2. Exports to Excel with helpful formatting
3. You fix candidate names, parties, offices manually
4. Converts back to CSV
5. Validates the results

Usage:
    # Step 1: Extract and prepare for editing
    py scripts/hybrid_workflow.py extract "data/2020 General Election.pdf"
    
    # Step 2: Edit the Excel file that opens
    
    # Step 3: Convert back to CSV
    py scripts/hybrid_workflow.py finalize "data/20201103__ky__general__county_EDIT.xlsx"
"""

import sys
import os
from pathlib import Path
import re
import pandas as pd
import logging

try:
    import tabula
    TABULA_AVAILABLE = True
except ImportError:
    TABULA_AVAILABLE = False

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

from pdf_utils import (
    clean_county_name, clean_votes, extract_party,
    get_election_date, validate_extraction_result
)


def extract_for_editing(pdf_path):
    """
    Extract data from PDF and prepare Excel file for manual editing.
    """
    pdf_path = Path(pdf_path)
    
    if not pdf_path.exists():
        logger.error(f"File not found: {pdf_path}")
        return None
    
    logger.info(f"\n{'='*70}")
    logger.info(f"STEP 1: Extracting from {pdf_path.name}")
    logger.info(f"{'='*70}\n")
    
    # Extract year
    year_match = re.search(r'(20\d{2})', pdf_path.name)
    year = year_match.group(1) if year_match else "unknown"
    
    # Try extraction with tabula
    results = []
    
    if TABULA_AVAILABLE:
        logger.info("Extracting tables...")
        try:
            tables = tabula.read_pdf(
                str(pdf_path),
                pages='all',
                multiple_tables=True,
                stream=True
            )
            
            if tables:
                logger.info(f"  Found {len(tables)} tables")
                
                for table_idx, df in enumerate(tables):
                    if df is None or df.empty or df.shape[1] < 2:
                        continue
                    
                    # Get county column (usually first)
                    county_col = df.columns[0]
                    
                    # Process each row
                    for idx, row in df.iterrows():
                        county_text = str(row[county_col])
                        county = clean_county_name(county_text)
                        
                        if not county:
                            continue
                        
                        # Process vote columns
                        for col in df.columns[1:]:
                            candidate = str(col).strip()
                            votes = clean_votes(row[col])
                            
                            results.append({
                                'county': county,
                                'office': '',  # TO FILL
                                'district': '',
                                'candidate': candidate,  # TO CORRECT
                                'party': '',  # TO FILL
                                'votes': votes
                            })
        except Exception as e:
            logger.warning(f"Tabula extraction had issues: {e}")
    
    # If no results, create template with counties
    if not results:
        logger.info("Creating county template for manual entry...")
        from pdf_utils import KY_COUNTIES
        
        for county in sorted(KY_COUNTIES):
            results.append({
                'county': county.title(),
                'office': '',
                'district': '',
                'candidate': 'ENTER_CANDIDATE_NAME',
                'party': '',
                'votes': 0
            })
    
    # Create DataFrame
    df = pd.DataFrame(results)
    
    logger.info(f"  Prepared {len(df)} rows for editing")
    
    # Add helper columns with instructions
    df.insert(0, 'INSTRUCTIONS →', [
        '← Edit these cells' if i == 0 else '' for i in range(len(df))
    ])
    
    # Create output filename
    election_date = get_election_date(year)
    output_excel = f"data/{election_date}__ky__general__county_EDIT.xlsx"
    
    # Save to Excel with formatting
    logger.info(f"\nSaving to Excel: {output_excel}")
    
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        # Add instructions sheet
        instructions = pd.DataFrame({
            'Step': [1, 2, 3, 4, 5],
            'What to Do': [
                'Review CANDIDATE column - fix names (remove generic "Candidate 1", etc.)',
                'Fill in PARTY column (REP, DEM, LIB, IND, GRN, etc.)',
                'Fill in OFFICE column (President, U.S. Senate, Governor, etc.)',
                'Verify VOTES are correct (check against PDF)',
                'Save and close Excel, then run: py scripts/hybrid_workflow.py finalize "THIS_FILE.xlsx"'
            ],
            'Example': [
                'Change "Candidate 1" to "Donald Trump"',
                'REP, DEM, LIB, IND',
                'President, U.S. Senate, U.S. House',
                'Compare totals with PDF',
                'Final step to create CSV'
            ]
        })
        instructions.to_excel(writer, sheet_name='INSTRUCTIONS_READ_FIRST', index=False)
        
        # Format Data sheet
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Highlight columns that need editing
        edit_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # Get column indices (office, candidate, party)
        col_letters = {
            'office': None,
            'candidate': None,
            'party': None
        }
        
        for idx, cell in enumerate(worksheet[1], 1):
            col_name = str(cell.value).lower()
            if 'office' in col_name:
                col_letters['office'] = cell.column_letter
            elif 'candidate' in col_name:
                col_letters['candidate'] = cell.column_letter
            elif 'party' in col_name:
                col_letters['party'] = cell.column_letter
        
        # Highlight edit columns
        for col_name, col_letter in col_letters.items():
            if col_letter:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row}']
                    cell.fill = edit_fill
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 15  # Instructions
        worksheet.column_dimensions['B'].width = 15  # County
        worksheet.column_dimensions['C'].width = 20  # Office
        worksheet.column_dimensions['D'].width = 10  # District
        worksheet.column_dimensions['E'].width = 30  # Candidate
        worksheet.column_dimensions['F'].width = 10  # Party
        worksheet.column_dimensions['G'].width = 12  # Votes
        
        # Format instructions sheet
        inst_ws = writer.sheets['INSTRUCTIONS_READ_FIRST']
        for cell in inst_ws[1]:
            cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            cell.font = Font(bold=True)
        
        inst_ws.column_dimensions['A'].width = 8
        inst_ws.column_dimensions['B'].width = 80
        inst_ws.column_dimensions['C'].width = 40
    
    logger.info(f"\n{'='*70}")
    logger.info("✓ READY FOR EDITING!")
    logger.info(f"{'='*70}")
    logger.info(f"\nNext steps:")
    logger.info(f"  1. Opening Excel file: {output_excel}")
    logger.info(f"  2. Read the INSTRUCTIONS sheet")
    logger.info(f"  3. Edit the Data sheet:")
    logger.info(f"     - Fix candidate names")
    logger.info(f"     - Add party affiliations")
    logger.info(f"     - Add office names")
    logger.info(f"     - Verify vote counts")
    logger.info(f"  4. Save and close Excel")
    logger.info(f"  5. Run: py scripts/hybrid_workflow.py finalize \"{output_excel}\"")
    logger.info(f"\n{'='*70}\n")
    
    # Try to open Excel automatically
    try:
        import subprocess
        subprocess.Popen(['start', '', output_excel], shell=True)
        logger.info("Opening Excel file...")
    except:
        logger.info(f"Please open manually: {output_excel}")
    
    return output_excel


def finalize_extraction(excel_path):
    """
    Convert edited Excel file back to CSV and validate.
    """
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        logger.error(f"File not found: {excel_path}")
        return None
    
    logger.info(f"\n{'='*70}")
    logger.info(f"STEP 2: Finalizing {excel_path.name}")
    logger.info(f"{'='*70}\n")
    
    # Read Excel
    logger.info("Reading edited Excel file...")
    try:
        df = pd.read_excel(excel_path, sheet_name='Data')
    except Exception as e:
        logger.error(f"Error reading Excel: {e}")
        return None
    
    # Remove instruction column if present
    if 'INSTRUCTIONS' in df.columns[0] or 'INSTRUCTIONS →' in str(df.columns[0]):
        df = df.iloc[:, 1:]
    
    # Ensure required columns
    required = ['county', 'office', 'district', 'candidate', 'party', 'votes']
    for col in required:
        if col not in df.columns:
            logger.error(f"Missing required column: {col}")
            return None
    
    # Add OpenElections format columns if missing
    oe_cols = ['election_day', 'absentee', 'av_counting_boards', 'early_voting', 
               'mail', 'provisional', 'pre_process_absentee']
    for col in oe_cols:
        if col not in df.columns:
            df[col] = ''
    
    # Clean data
    df['votes'] = pd.to_numeric(df['votes'], errors='coerce').fillna(0).astype(int)
    df['county'] = df['county'].fillna('')
    df['candidate'] = df['candidate'].fillna('')
    
    # Remove empty rows
    df = df[df['county'].str.strip() != '']
    df = df[df['candidate'].str.strip() != '']
    
    logger.info(f"  Processed {len(df)} rows")
    
    # Validate
    logger.info("\nValidating data...")
    is_valid, warnings = validate_extraction_result(df)
    
    if warnings:
        logger.warning("Validation warnings:")
        for warning in warnings:
            logger.warning(f"  - {warning}")
    
    # Check for common issues
    issues = []
    
    # Check for generic candidate names
    generic_names = df[df['candidate'].str.contains('Candidate \d+|ENTER_', case=False, na=False)]
    if len(generic_names) > 0:
        issues.append(f"Found {len(generic_names)} rows with generic candidate names")
    
    # Check for missing parties
    missing_party = df[df['party'].str.strip() == '']
    if len(missing_party) > len(df) * 0.5:
        issues.append(f"{len(missing_party)} rows missing party information")
    
    # Check for missing offices
    missing_office = df[df['office'].str.strip() == '']
    if len(missing_office) > len(df) * 0.5:
        issues.append(f"{len(missing_office)} rows missing office information")
    
    if issues:
        logger.warning("\nData quality issues found:")
        for issue in issues:
            logger.warning(f"  ⚠️  {issue}")
        
        proceed = input("\nContinue anyway? (y/n): ").strip().lower()
        if proceed != 'y':
            logger.info("Cancelled. Please edit Excel file and try again.")
            return None
    
    # Generate output CSV filename
    csv_name = excel_path.stem.replace('_EDIT', '') + '.csv'
    csv_path = excel_path.parent / csv_name
    
    # Select columns in OpenElections order
    column_order = [
        'county', 'office', 'district', 'candidate', 'party', 'votes',
        'election_day', 'absentee', 'av_counting_boards', 'early_voting',
        'mail', 'provisional', 'pre_process_absentee'
    ]
    
    df_final = df[column_order]
    
    # Save to CSV
    df_final.to_csv(csv_path, index=False)
    
    logger.info(f"\n{'='*70}")
    logger.info("✓ SUCCESS!")
    logger.info(f"{'='*70}")
    logger.info(f"\n✓ Saved final CSV: {csv_path}")
    logger.info(f"  Records: {len(df_final)}")
    logger.info(f"  Counties: {df_final['county'].nunique()}")
    logger.info(f"  Candidates: {df_final['candidate'].nunique()}")
    logger.info(f"  Total votes: {df_final['votes'].sum():,}")
    
    # Show preview
    print("\nPreview (first 10 rows):")
    print(df_final.head(10).to_string(index=False))
    
    logger.info(f"\n{'='*70}")
    logger.info("Next step: Validate your data")
    logger.info(f"{'='*70}")
    logger.info(f"\nRun: py scripts/validate_extraction.py \"{csv_path}\"")
    logger.info(f"\n{'='*70}\n")
    
    return csv_path


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Command to run')
    
    # Extract command
    extract_parser = subparsers.add_parser('extract', help='Extract PDF to Excel for editing')
    extract_parser.add_argument('pdf_path', help='Path to PDF file')
    
    # Finalize command
    finalize_parser = subparsers.add_parser('finalize', help='Convert edited Excel to CSV')
    finalize_parser.add_argument('excel_path', help='Path to edited Excel file')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    if args.command == 'extract':
        result = extract_for_editing(args.pdf_path)
        if result:
            print(f"\n💡 TIP: View PDF and Excel side-by-side for easy editing!")
    
    elif args.command == 'finalize':
        result = finalize_extraction(args.excel_path)
        if result:
            print(f"\n🎉 All done! Your data is ready to use.")


if __name__ == '__main__':
    main()
